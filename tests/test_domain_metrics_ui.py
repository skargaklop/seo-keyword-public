# Test: Domain metrics UI rendering and Excel export (Plan 14-02 Task 2)
# Purpose: Verify render_serp_domain_metrics function and Excel export integration
# LINKS: PLAN 14-02 Task 2
# MODULE_CONTRACT: tests/test_domain_metrics_ui
# Purpose: Verify domain metrics rendering and export
# Rationale: Tests UI rendering with mock session state and multi-sheet export with domain sheet
# Dependencies: components.results, utils.seo_math_analysis, utils.excel_exporter
# Exports: pytest tests
# Verification: python -m pytest tests/test_domain_metrics_ui.py -q
# CHANGE_SUMMARY: Added domain metrics UI tests — covers render_serp_domain_metrics, Excel export sheet, session state lifecycle

import io
from unittest.mock import patch

import pandas as pd

from utils.seo_math_analysis import DomainMetrics


# Purpose: Test render_serp_domain_metrics function.
class TestRenderSerpDomainMetrics:

    # Purpose: render_serp_domain_metrics must be importable from components.results.
    def test_render_function_is_importable(self):
        from components.results import render_serp_domain_metrics
        assert callable(render_serp_domain_metrics)

    # Purpose: render_serp_domain_metrics returns None.
    def test_render_returns_none(self):
        from components.results import render_serp_domain_metrics
        with patch.dict("streamlit.session_state", {"serp_domain_metrics": None}):
            result = render_serp_domain_metrics()
            assert result is None

    # Purpose: render_serp_domain_metrics handles None session state gracefully.
    def test_render_does_not_crash_with_none_session_state(self):
        from components.results import render_serp_domain_metrics
        with patch.dict("streamlit.session_state", {"serp_domain_metrics": None}):
            # Should not raise
            render_serp_domain_metrics()

    # Purpose: render_serp_domain_metrics handles empty list gracefully.
    def test_render_does_not_crash_with_empty_list(self):
        from components.results import render_serp_domain_metrics
        with patch.dict("streamlit.session_state", {"serp_domain_metrics": []}):
            render_serp_domain_metrics()

    # Purpose: render_serp_domain_metrics renders domain metrics from session state.
    def test_render_with_domain_metrics(self):
        from components.results import render_serp_domain_metrics
        metrics = [
            DomainMetrics(
                domain="example.com",
                avg_position=2.5,
                keyword_serp_count=3,
                total_keyword_serps=5,
                result_count=10,
                total_results=50,
            ),
        ]
        with patch.dict("streamlit.session_state", {"serp_domain_metrics": metrics}):
            with patch("streamlit.subheader") as mock_subheader, \
                 patch("streamlit.dataframe") as mock_dataframe:
                render_serp_domain_metrics()
                # Should render a subheader and dataframe
                mock_subheader.assert_called()
                mock_dataframe.assert_called()

    # Purpose: Domain metrics display should format avg_position to 1 decimal.
    def test_render_formats_avg_position_to_one_decimal(self):
        from components.results import render_serp_domain_metrics
        from config.i18n import t
        metrics = [
            DomainMetrics(
                domain="example.com",
                avg_position=2.5678,
                keyword_serp_count=3,
                total_keyword_serps=5,
                result_count=10,
                total_results=50,
            ),
        ]
        with patch.dict("streamlit.session_state", {"serp_domain_metrics": metrics}):
            with patch("streamlit.subheader"), \
                 patch("streamlit.dataframe") as mock_dataframe:
                render_serp_domain_metrics()
                # Verify the dataframe was called with rounded data
                call_args = mock_dataframe.call_args
                df_arg = call_args[0][0] if call_args[0] else call_args[1].get("data")
                if df_arg is not None:
                    # Avg position column is now i18n-localized
                    avg_col = t("serp_domain_avg_position")
                    assert df_arg.iloc[0][avg_col] == 2.6


# Purpose: Test domain metrics export in multi-sheet Excel.
class TestDomainMetricsExcelExport:

    # Purpose: Multi-sheet export should include domain metrics sheet when data exists.
    def test_domain_metrics_sheet_in_export(self):
        from utils.excel_exporter import ExcelExporter

        metrics = [
            DomainMetrics(
                domain="example.com",
                avg_position=2.5,
                keyword_serp_count=3,
                total_keyword_serps=5,
                result_count=10,
                total_results=50,
            ),
        ]

        # Build domain metrics DataFrame as the export would
        domain_rows = []
        for dm in metrics:
            domain_rows.append({
                "Domain": dm.domain,
                "Avg Position": round(dm.avg_position, 1),
                "Keyword SERPs": f"{dm.keyword_serp_count}/{dm.total_keyword_serps}",
                "Total Keywords": dm.total_keyword_serps,
                "Result Count": dm.result_count,
                "Total Results": dm.total_results,
            })

        sheets = {
            "SERP Results": pd.DataFrame({"Keyword": ["test"], "Position": [1]}),
            "SERP Domains": pd.DataFrame(domain_rows),
        }

        buffer = io.BytesIO()
        result = ExcelExporter.export_multi_sheet_to_buffer(sheets, buffer)
        assert result is True
        assert buffer.tell() == 0  # seeked back to start

        # Verify the workbook has the domain sheet
        from openpyxl import load_workbook
        buffer.seek(0)
        wb = load_workbook(buffer)
        sheet_names = wb.sheetnames
        assert "SERP Domains" in sheet_names

    # Purpose: Domain metrics export uses existing export_multi_sheet_to_buffer API.
    def test_domain_metrics_export_uses_existing_multi_sheet_api(self):
        from utils.excel_exporter import ExcelExporter

        # Verify the API exists and has correct signature
        import inspect
        sig = inspect.signature(ExcelExporter.export_multi_sheet_to_buffer)
        params = list(sig.parameters.keys())
        assert "sheets" in params
        assert "buffer" in params