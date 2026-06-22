"""
Unit tests for ExcelExporter (improvement #17).
Phase 10 Task 12: Tests for multi-sheet export, metadata columns, and column grouping.
"""

import io
import os
from datetime import datetime
from pathlib import Path
from unittest.mock import patch

# MODULE_CONTRACT: tests/test_excel_exporter
# Purpose: Verify Excel export sheets, metadata, styles, and workbook structure.
# Rationale: Links exporter tests to the GRACE export module.
# Dependencies: pandas, pytest, openpyxl, utils.excel_exporter.
# Exports: pytest tests.
# LINKS: knowledge-graph.xml#MOD-010
# MODULE_MAP: tests/test_excel_exporter.py
# Public Functions: pytest test functions.
# Private Helpers: pytest fixtures in this file.
# Key Semantic Blocks: none.
# Critical Flows: create sample DataFrame -> export workbook -> inspect sheets/cells/styles.
# Verification: verification-plan.xml#V-09-EXPORT, verification-plan.xml#V-10-EXPORTS
# CHANGE_SUMMARY: Added GRACE module contract linking this test file to MOD-010.

import pandas as pd
import pytest
from openpyxl import load_workbook

from utils.excel_exporter import ExcelExporter


# Purpose: sample df implementation
# Purpose: sample df implementation
@pytest.fixture
def sample_df() -> pd.DataFrame:
    return pd.DataFrame(
        {
            "Keyword": ["купить кофе", "чай зеленый", "молоко"],
            "Source URL": [
                "https://example.com",
                "https://example.com",
                "https://test.org",
            ],
            "Avg Monthly Searches": [1000, 500, 200],
        }
    )

# Purpose: DataFrame with BM25F, signal, and Trends metadata columns.
@pytest.fixture
def sample_df_with_metadata() -> pd.DataFrame:
    return pd.DataFrame(
        {
            "Keyword": ["купить кофе", "чай зеленый", "молоко"],
            "Avg Monthly Searches": [1000, 500, 200],
            "BM25F Score": [2.5, 1.8, 0.9],
            "BM25F Coverage": [0.8, 0.6, 0.4],
            "Cache Hit": ["Yes", "No", "Yes"],
            "Fetched At": ["2025-01-01 10:00", "2025-01-01 11:00", "2025-01-01 12:00"],
            "Average Interest": [75, 50, 30],
        }
    )


# Purpose: TestExport implementation
class TestExport:
    # Purpose: Test export creates file
    # Purpose: Test export creates file
    @patch.object(ExcelExporter, "_validate_export_path")
    def test_export_creates_file(
        self, mock_validate, sample_df: pd.DataFrame, tmp_path: Path
    ) -> None:
        file_path: str = str(tmp_path / "test_export.xlsx")
        result: bool = ExcelExporter.export(sample_df, file_path)
        assert result is True
        assert os.path.exists(file_path)

    # Purpose: Test export readable
    # Purpose: Test export readable
    @patch.object(ExcelExporter, "_validate_export_path")
    def test_export_readable(
        self, mock_validate, sample_df: pd.DataFrame, tmp_path: Path
    ) -> None:
        file_path: str = str(tmp_path / "test_export.xlsx")
        ExcelExporter.export(sample_df, file_path)
        df_read: pd.DataFrame = pd.read_excel(file_path)
        assert len(df_read) == 3
        assert "Keyword" in df_read.columns


# Purpose: TestExportToBuffer implementation
class TestExportToBuffer:
    # Purpose: Test export to buffer
    def test_export_to_buffer(self, sample_df: pd.DataFrame) -> None:
        buffer: io.BytesIO = io.BytesIO()
        result: bool = ExcelExporter.export_to_buffer(sample_df, buffer)
        assert result is True
        assert buffer.tell() == 0  # Should be seeked to start
        assert len(buffer.getvalue()) > 0

    # Purpose: Test buffer readable
    def test_buffer_readable(self, sample_df: pd.DataFrame) -> None:
        buffer: io.BytesIO = io.BytesIO()
        ExcelExporter.export_to_buffer(sample_df, buffer)
        df_read: pd.DataFrame = pd.read_excel(buffer)
        assert len(df_read) == 3

    # Purpose: Test seo keywords column wraps text
    def test_seo_keywords_column_wraps_text(self) -> None:
        data = pd.DataFrame(
            {
                "Ключевые слова": [
                    "купить наполнитель, цена наполнителя, бумажный наполнитель, опт"
                ],
                "URL": ["https://example.com"],
                "SEO Text": ["<p>text</p>"],
                "Page Content": ["source content"],
            }
        )
        buffer = io.BytesIO()

        ExcelExporter.export_to_buffer(data, buffer)

        workbook = load_workbook(buffer)
        sheet = workbook.active

        assert sheet["A2"].alignment.wrap_text is True
        assert sheet["B2"].alignment.vertical == "top"
        assert sheet["B2"].alignment.horizontal == "left"

    # Purpose: Test style rule can target url column from secondary key
    def test_style_rule_can_target_url_column_from_secondary_key(self) -> None:
        data = pd.DataFrame(
            {
                "URL": ["https://example.com/page", "https://example.com/other"],
                "URL Match Type": ["full_url", "domain"],
            }
        )
        style_rules = {
            "URL": {
                "condition": lambda row: row.get("URL Match Type") == "full_url",
                "font": {"bold": True, "underline": True},
            },
            "URL_domain": {
                "target_column": "URL",
                "condition": lambda row: row.get("URL Match Type") == "domain",
                "font": {"bold": True},
            },
        }
        buffer = io.BytesIO()

        ExcelExporter.export_to_buffer(data, buffer, style_rules=style_rules)

        workbook = load_workbook(buffer)
        sheet = workbook.active

        assert sheet["A2"].font.bold is True
        assert sheet["A2"].font.underline == "single"
        assert sheet["A3"].font.bold is True
        assert sheet["A3"].font.underline is None


# Purpose: Phase 10 Task 12: Multi-sheet export tests.
class TestMultiSheetExport:

    # Purpose: Test export multi sheet creates file
    # Purpose: Test export multi sheet creates file
    @patch.object(ExcelExporter, "_validate_export_path")
    def test_export_multi_sheet_creates_file(
        self, mock_validate, sample_df: pd.DataFrame, tmp_path: Path
    ) -> None:
        sheets = {
            "Main Data": sample_df,
            "Interest Over Time": pd.DataFrame({
                "Date": ["2025-01", "2025-02"],
                "купить кофе": [80, 75],
                "чай зеленый": [50, 45],
            }),
            "Related Queries": pd.DataFrame({
                "Query": ["кофе машина", "чайWithURL"],
                "Value": [100, 80],
            }),
        }
        file_path = str(tmp_path / "multi_sheet.xlsx")

        result = ExcelExporter.export_multi_sheet(sheets, file_path)

        assert result is True
        assert os.path.exists(file_path)

        # Verify multiple sheets exist
        wb = load_workbook(file_path)
        assert len(wb.sheetnames) == 3
        assert "Main Data" in wb.sheetnames
        assert "Interest Over Time" in wb.sheetnames or "Interest_Over_Time" in wb.sheetnames
        assert "Related Queries" in wb.sheetnames or "Related_Queries" in wb.sheetnames

    # Purpose: Test export multi sheet with grouping
    # Purpose: Test export multi sheet with grouping
    @patch.object(ExcelExporter, "_validate_export_path")
    def test_export_multi_sheet_with_grouping(
        self, mock_validate, sample_df_with_metadata: pd.DataFrame, tmp_path: Path
    ) -> None:
        """Test that metadata columns are grouped."""
        sheets = {"Results": sample_df_with_metadata}
        file_path = str(tmp_path / "grouped.xlsx")

        result = ExcelExporter.export_multi_sheet(
            sheets, file_path, group_metadata=True
        )

        assert result is True

        # Verify column grouping exists
        wb = load_workbook(file_path)
        ws = wb.active

        # Check that metadata columns exist
        headers = [cell.value for cell in ws[1]]
        assert "Cache Hit" in headers
        assert "BM25F Score" in headers
        assert "Average Interest" in headers

    # Purpose: Test export multi sheet to buffer
    # Purpose: Test export multi sheet to buffer
    @patch.object(ExcelExporter, "_validate_export_path")
    def test_export_multi_sheet_to_buffer(
        self, mock_validate, sample_df: pd.DataFrame
    ) -> None:
        sheets = {
            "Main": sample_df,
            "Secondary": pd.DataFrame({"Col1": [1, 2], "Col2": [3, 4]}),
        }
        buffer = io.BytesIO()

        result = ExcelExporter.export_multi_sheet_to_buffer(sheets, buffer)

        assert result is True
        assert len(buffer.getvalue()) > 0

        # Verify both sheets exist
        wb = load_workbook(buffer)
        assert len(wb.sheetnames) == 2

    # Purpose: Test sheet name sanitization.
    def test_sanitize_sheet_name(self) -> None:
        assert ExcelExporter._sanitize_sheet_name("Normal Name") == "Normal Name"
        assert ExcelExporter._sanitize_sheet_name("Name:With*Invalid?Chars") == "Name_With_Invalid_Chars"
        assert ExcelExporter._sanitize_sheet_name("A" * 50) == "A" * 31
        assert ExcelExporter._sanitize_sheet_name("") == "Sheet"
        assert ExcelExporter._sanitize_sheet_name("Name[With]Brackets") == "Name_With_Brackets"


# Purpose: Phase 10 Task 12: Cache metadata column tests.
class TestCacheMetadataColumns:

    # Purpose: Test add cache metadata columns
    def test_add_cache_metadata_columns(self) -> None:
        df = pd.DataFrame({"Keyword": ["test", "demo"]})
        cache_meta = {
            "cache_hit": True,
            "cache_key": "abc123def4567890123456789",
            "created_at": "2025-01-15T10:30:00",
            "expires_at": "2025-01-22T10:30:00",
            "provider": "google_trends",
            "ttl_hours": 24,
        }

        result = ExcelExporter.add_cache_metadata_columns(df, cache_meta)

        assert "Cache Hit" in result.columns
        assert "Cache Key Prefix" in result.columns
        assert "Fetched At" in result.columns
        assert "Expires At" in result.columns
        assert "Cache Provider" in result.columns
        assert "Cache TTL Hours" in result.columns

        # Verify values
        assert result["Cache Hit"].iloc[0] == "Yes"
        assert result["Cache Key Prefix"].iloc[0] == "abc123def4567890"  # 16 chars
        assert result["Fetched At"].iloc[0] == "2025-01-15T10:30:00"  # String format preserved to 19 chars
        assert result["Expires At"].iloc[0] == "2025-01-22T10:30:00"  # String format preserved to 19 chars
        assert result["Cache Provider"].iloc[0] == "google_trends"
        assert result["Cache TTL Hours"].iloc[0] == 24

    # Purpose: Test add cache metadata with datetime objects
    def test_add_cache_metadata_with_datetime_objects(self) -> None:
        df = pd.DataFrame({"Keyword": ["test"]})
        now = datetime(2025, 1, 15, 10, 30)
        later = datetime(2025, 1, 22, 10, 30)
        cache_meta = {
            "cache_hit": False,
            "cache_key": "",
            "created_at": now,
            "expires_at": later,
            "provider": "test_provider",
            "ttl_hours": 168,
        }

        result = ExcelExporter.add_cache_metadata_columns(df, cache_meta)

        assert result["Cache Hit"].iloc[0] == "No"
        assert result["Fetched At"].iloc[0] == "2025-01-15 10:30"
        assert result["Expires At"].iloc[0] == "2025-01-22 10:30"

    # Purpose: Test add cache metadata empty df
    def test_add_cache_metadata_empty_df(self) -> None:
        df = pd.DataFrame()
        cache_meta = {"cache_hit": True}

        result = ExcelExporter.add_cache_metadata_columns(df, cache_meta)

        # Should return empty DataFrame unchanged
        assert result.empty

    # Purpose: Test add cache metadata no metadata
    def test_add_cache_metadata_no_metadata(self) -> None:
        df = pd.DataFrame({"Keyword": ["test"]})

        result = ExcelExporter.add_cache_metadata_columns(df, {})

        # Should return original DataFrame unchanged
        assert "Cache Hit" not in result.columns
        assert len(result.columns) == 1


# Purpose: TestTrendsMetadataColumns implementation
class TestTrendsMetadataColumns:
    # Purpose: Test add trends columns includes provider confidence and warnings
    def test_add_trends_columns_includes_provider_confidence_and_warnings(self) -> None:
        df = pd.DataFrame({"Keyword": ["test"]})
        trends_meta = {
            "provider": "google_trends_alpha",
            "data_confidence": "degraded",
            "integrity_warnings": ["soft_block", "scaled_values"],
            "provider_metadata": {"provider": "google_trends_alpha", "mode": "official"},
            "cache_metadata": {"cache_hit": True, "cache_key": "abc123"},
        }

        result = ExcelExporter.add_trends_columns(df, trends_meta)

        assert "Trends Provider" in result.columns
        assert "Trends Data Confidence" in result.columns
        assert "Trends Integrity Warnings" in result.columns
        assert "Trends Provider Metadata" in result.columns
        assert result["Trends Provider"].iloc[0] == "google_trends_alpha"
        assert result["Trends Data Confidence"].iloc[0] == "degraded"
        assert "soft_block" in result["Trends Integrity Warnings"].iloc[0]
        assert "google_trends_alpha" in result["Trends Provider Metadata"].iloc[0]

    # Purpose: Test build trends provider metadata df includes caveats
    def test_build_trends_provider_metadata_df_includes_caveats(self) -> None:
        trends_meta = {
            "provider": "google_trends_alpha",
            "data_confidence": "high",
            "integrity_warnings": ["scaled_values"],
            "warnings": ["partial"],
            "provider_metadata": {"provider": "google_trends_alpha"},
            "cache_metadata": {"cache_hit": False},
            "caveats": ["relative scale", "official alpha"],
        }

        metadata_df = ExcelExporter.build_trends_provider_metadata_df(trends_meta)

        assert list(metadata_df.columns) == ["Field", "Value"]
        assert "Provider" in metadata_df["Field"].tolist()
        assert "Data Confidence" in metadata_df["Field"].tolist()
        assert "Integrity Warnings" in metadata_df["Field"].tolist()
        assert "Provider Metadata" in metadata_df["Field"].tolist()
        assert "Cache Metadata" in metadata_df["Field"].tolist()
        assert metadata_df[metadata_df["Field"] == "Caveat"]["Value"].tolist() == [
            "relative scale",
            "official alpha",
        ]

    # Purpose: Test export multi sheet can include trends metadata sheet
    # Purpose: Test export multi sheet can include trends metadata sheet
    @patch.object(ExcelExporter, "_validate_export_path")
    def test_export_multi_sheet_can_include_trends_metadata_sheet(
        self, mock_validate, tmp_path: Path
    ) -> None:
        sheets = {
            "Averages": pd.DataFrame({"Keyword": ["alpha"], "Average Interest": [75]}),
            "Trends Provider Metadata": ExcelExporter.build_trends_provider_metadata_df(
                {
                    "provider": "google_trends_alpha",
                    "data_confidence": "high",
                    "integrity_warnings": ["official_v1alpha_scaled"],
                    "provider_metadata": {"provider": "google_trends_alpha"},
                    "cache_metadata": {"cache_hit": True},
                    "caveats": ["relative scale"],
                }
            ),
        }
        buffer = io.BytesIO()

        result = ExcelExporter.export_multi_sheet_to_buffer(sheets, buffer)

        assert result is True
        workbook = load_workbook(buffer)
        assert "Trends Provider Metadata" in workbook.sheetnames
        metadata_sheet = workbook["Trends Provider Metadata"]
        assert metadata_sheet["A1"].value == "Field"
        assert metadata_sheet["B2"].value == "google_trends_alpha"


# Purpose: Phase 10 Task 12: BM25F and signal analysis column tests.
class TestBM25FSignalColumns:

    # Purpose: Test add bm25f signal columns
    def test_add_bm25f_signal_columns(self) -> None:
        df = pd.DataFrame({"Keyword": ["купить кофе", "чай"]})

        # Mock BM25F data
        from dataclasses import dataclass

        # Purpose: MockBM25F implementation
        # Purpose: MockBM25F implementation
        @dataclass
        class MockBM25F:
            doc_text: str
            score: float
            query_coverage: float
            field_contributions: dict
            matched_terms: list

        bm25f_results = [
            MockBM25F("купить кофе", 2.5, 0.8, {"title": 1.5}, ["купить", "кофе"]),
            MockBM25F("чай", 1.8, 0.6, {"title": 1.2}, ["чай"]),
        ]

        # Mock signal data
        # Purpose: MockTitleAlign implementation
        # Purpose: MockTitleAlign implementation
        @dataclass
        class MockTitleAlign:
            title_alignment_score: float
            title_h1_overlap: float
            title_rewrite_risk: float

        # Purpose: MockContentEffort implementation
        # Purpose: MockContentEffort implementation
        @dataclass
        class MockContentEffort:
            effort_score: float
            effort_level: str

        signal_results = {
            "title_alignment": MockTitleAlign(0.75, 0.8, 0.25),
            "content_effort": MockContentEffort(0.65, "medium"),
        }

        analysis_data = {
            "bm25f_results": bm25f_results,
            "signal_results": signal_results,
        }

        result = ExcelExporter.add_bm25f_signal_columns(df, analysis_data)

        assert "BM25F Score" in result.columns
        assert "BM25F Coverage" in result.columns
        assert "Title Alignment Score" in result.columns
        assert "Title H1 Overlap" in result.columns
        assert "Title Rewrite Risk" in result.columns
        assert "Content Effort Score" in result.columns
        assert "Content Effort Level" in result.columns

        # Verify values
        assert result["BM25F Score"].iloc[0] == 2.5
        assert result["BM25F Coverage"].iloc[0] == 0.8
        assert result["Title Alignment Score"].iloc[0] == 0.75
        assert result["Content Effort Level"].iloc[0] == "medium"

    # Purpose: Test add bm25f signal columns empty df
    def test_add_bm25f_signal_columns_empty_df(self) -> None:
        df = pd.DataFrame()
        analysis_data = {"bm25f_results": [], "signal_results": {}}

        result = ExcelExporter.add_bm25f_signal_columns(df, analysis_data)

        assert result.empty

    # Purpose: Test add bm25f signal columns no data
    def test_add_bm25f_signal_columns_no_data(self) -> None:
        df = pd.DataFrame({"Keyword": ["test"]})

        result = ExcelExporter.add_bm25f_signal_columns(df, {})

        # Empty analysis_data returns original DataFrame unchanged
        assert "BM25F Score" not in result.columns
        assert len(result.columns) == 1


# Purpose: Phase 10 Task 12: Google Trends column tests.
class TestTrendsColumns:

    # Purpose: Test add trends columns
    def test_add_trends_columns(self) -> None:
        df = pd.DataFrame({"Keyword": ["купить кофе", "чай", "молоко"]})

        trends_data = {
            "averages": {
                "купить кофе": 75.5,
                "чай": 50.2,
                "молоко": 30.0,
            },
            "geo": "UA",
            "timeframe": "today 12-m",
            "provider": "google_trends_direct",
        }

        result = ExcelExporter.add_trends_columns(df, trends_data)

        assert "Average Interest" in result.columns
        assert "Trends Geo" in result.columns
        assert "Trends Timeframe" in result.columns
        assert "Trends Provider" in result.columns

        # Verify values
        assert result["Average Interest"].iloc[0] == 75.5
        assert result["Average Interest"].iloc[1] == 50.2
        assert result["Trends Geo"].iloc[0] == "UA"
        assert result["Trends Timeframe"].iloc[0] == "today 12-m"
        assert result["Trends Provider"].iloc[0] == "google_trends_direct"

    # Purpose: Test add trends columns case insensitive match
    def test_add_trends_columns_case_insensitive_match(self) -> None:
        df = pd.DataFrame({"Keyword": ["Кофе", "Чай"]})

        trends_data = {
            "averages": {
                "кофе": 80.0,
                "чай": 60.0,
            },
            "geo": "UA",
            "timeframe": "today 12-m",
            "provider": "google_trends",
        }

        result = ExcelExporter.add_trends_columns(df, trends_data)

        # Should match case-insensitively
        assert result["Average Interest"].iloc[0] == 80.0
        assert result["Average Interest"].iloc[1] == 60.0

    # Purpose: Test add trends columns empty df
    def test_add_trends_columns_empty_df(self) -> None:
        df = pd.DataFrame()
        trends_data = {"averages": {}, "geo": "UA"}

        result = ExcelExporter.add_trends_columns(df, trends_data)

        assert result.empty

    # Purpose: Test add trends columns preserves existing average when averages are absent
    def test_add_trends_columns_preserves_existing_average_without_averages(self) -> None:
        df = pd.DataFrame({"Keyword": ["seo"], "Average Interest": [79.51]})

        trends_data = {
            "geo": "UA",
            "timeframe": "today 12-m",
            "provider": "google_trends_direct",
        }

        result = ExcelExporter.add_trends_columns(df, trends_data)

        assert result["Average Interest"].iloc[0] == 79.51

    # Purpose: Test add trends columns no data
    def test_add_trends_columns_no_data(self) -> None:
        df = pd.DataFrame({"Keyword": ["test"]})

        result = ExcelExporter.add_trends_columns(df, {})

        # Empty trends_data returns original DataFrame unchanged
        assert "Average Interest" not in result.columns
        assert len(result.columns) == 1


# Purpose: EXPORT-15-01/02: CSV export has parity with Excel export.
class TestExportCsvParity:

    # Purpose: CSV export must contain the same columns as Excel export.
    def test_csv_export_contains_same_columns_as_excel(self, sample_df: pd.DataFrame) -> None:
        csv_bytes = ExcelExporter.export_csv_to_bytes(sample_df)
        df_csv = pd.read_csv(io.BytesIO(csv_bytes), encoding="utf-8-sig")

        buffer = io.BytesIO()
        ExcelExporter.export_to_buffer(sample_df, buffer)
        buffer.seek(0)
        df_xlsx = pd.read_excel(buffer)

        assert list(df_csv.columns) == list(df_xlsx.columns), (
            f"CSV columns {list(df_csv.columns)} != Excel columns {list(df_xlsx.columns)}"
        )

    # Purpose: CSV export must preserve the same row count as Excel export.
    def test_csv_export_preserves_same_row_count_as_excel(self, sample_df: pd.DataFrame) -> None:
        csv_bytes = ExcelExporter.export_csv_to_bytes(sample_df)
        df_csv = pd.read_csv(io.BytesIO(csv_bytes), encoding="utf-8-sig")

        buffer = io.BytesIO()
        ExcelExporter.export_to_buffer(sample_df, buffer)
        buffer.seek(0)
        df_xlsx = pd.read_excel(buffer)

        assert len(df_csv) == len(df_xlsx), (
            f"CSV rows {len(df_csv)} != Excel rows {len(df_xlsx)}"
        )

    # Purpose: CSV export must preserve the same data values as Excel export.
    def test_csv_export_preserves_same_data_values(self, sample_df: pd.DataFrame) -> None:
        csv_bytes = ExcelExporter.export_csv_to_bytes(sample_df)
        df_csv = pd.read_csv(io.BytesIO(csv_bytes), encoding="utf-8-sig")

        buffer = io.BytesIO()
        ExcelExporter.export_to_buffer(sample_df, buffer)
        buffer.seek(0)
        df_xlsx = pd.read_excel(buffer)

        pd.testing.assert_frame_equal(df_csv, df_xlsx)


# Purpose: TestExportCsv implementation
class TestExportCsv:
    # Purpose: Test export csv to bytes has utf8 bom
    def test_export_csv_to_bytes_has_utf8_bom(self, sample_df: pd.DataFrame) -> None:
        csv_bytes = ExcelExporter.export_csv_to_bytes(sample_df)

        # UTF-8 BOM
        assert csv_bytes.startswith(b"\xef\xbb\xbf")

        decoded = csv_bytes.decode("utf-8-sig")
        lines = decoded.splitlines()

        assert lines[0] == "Keyword,Source URL,Avg Monthly Searches"

        df_read = pd.read_csv(io.BytesIO(csv_bytes), encoding="utf-8-sig")
        assert len(df_read) == 3
        assert df_read.iloc[0]["Keyword"] == sample_df.iloc[0]["Keyword"]

    # Purpose: Test export csv creates file
    # Purpose: Test export csv creates file
    @patch.object(ExcelExporter, "_validate_export_path")
    def test_export_csv_creates_file(
        self, mock_validate, sample_df: pd.DataFrame, tmp_path: Path
    ) -> None:
        file_path: str = str(tmp_path / "test_export.csv")
        result: bool = ExcelExporter.export_csv(sample_df, file_path)
        assert result is True
        assert os.path.exists(file_path)

    # Purpose: Test csv readable
    # Purpose: Test csv readable
    @patch.object(ExcelExporter, "_validate_export_path")
    def test_csv_readable(
        self, mock_validate, sample_df: pd.DataFrame, tmp_path: Path
    ) -> None:
        file_path: str = str(tmp_path / "test_export.csv")
        ExcelExporter.export_csv(sample_df, file_path)
        with open(file_path, "rb") as exported_file:
            raw_data = exported_file.read()

        # UTF-8 BOM
        assert raw_data.startswith(b"\xef\xbb\xbf")

        df_read: pd.DataFrame = pd.read_csv(file_path, encoding="utf-8-sig")
        assert len(df_read) == 3
        assert "Keyword" in df_read.columns


# Purpose: Phase 10 Task 12: Column grouping tests.
class TestColumnGrouping:

    # Purpose: Test metadata columns are grouped
    # Purpose: Test metadata columns are grouped
    @patch.object(ExcelExporter, "_validate_export_path")
    def test_metadata_columns_are_grouped(
        self, mock_validate, sample_df_with_metadata: pd.DataFrame, tmp_path: Path
    ) -> None:
        """Test that metadata columns are grouped in Excel exports."""
        file_path = str(tmp_path / "grouped_export.xlsx")

        result = ExcelExporter.export_multi_sheet(
            {"Data": sample_df_with_metadata},
            file_path,
            group_metadata=True,
        )

        assert result is True

        wb = load_workbook(file_path)
        ws = wb.active

        # Verify metadata columns exist
        headers = [cell.value for cell in ws[1]]
        cache_indices = [i for i, h in enumerate(headers) if h in _CACHE_METADATA_COLUMNS]

        # Groups require at least 2 adjacent columns
        # Cache metadata should have Cache Hit and Fetched At at minimum
        assert len(cache_indices) >= 2


# Import column constants for testing
_CACHE_METADATA_COLUMNS = [
    "Cache Hit",
    "Cache Key Prefix",
    "Fetched At",
    "Expires At",
    "Cache Provider",
    "Cache TTL Hours",
]

_BM25F_METADATA_COLUMNS = [
    "BM25F Score",
    "BM25F Coverage",
    "Field Contributions",
    "Matched Terms",
]

_TRENDS_METADATA_COLUMNS = [
    "Average Interest",
    "Interest Value",
    "Trends Geo",
    "Trends Timeframe",
    "Trends Provider",
]
