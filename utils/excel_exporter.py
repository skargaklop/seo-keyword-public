# MODULE_CONTRACT: utils/excel_exporter
# Purpose: Excel exporter module — exports DataFrames to formatted Excel files with multi-sheet support for BM25F, signals, Trends, and cache metadata.
# Rationale: Keep the module boundary explicit for GRACE adoption and review.
# Dependencies: io, pathlib, typing, pandas, openpyxl, openpyxl.styles, openpyxl.worksheet.worksheet, utils.logger
# Exports: ExcelExporter
# LINKS: requirements.xml#UC-001, requirements.xml#EXPT-10-01, development-plan.xml#MOD-010
# MODULE_MAP: utils/excel_exporter.py
# Public Functions: exported callables and classes defined in this module
# Private Helpers: internal helpers and private methods defined in this module
# Key Semantic Blocks: block_export_single_sheet, block_export_multi_sheet, block_excel_format_workbook, block_apply_style_rules, block_group_metadata_columns
# Critical Flows: preserve existing runtime behavior and integrations; support multi-sheet exports with metadata grouping
# Verification: python -m py_compile, python -m ruff check ., python -m pytest -q
# CHANGE_SUMMARY: Phase 10 Task 12: Added multi-sheet export support, BM25F/signal/Trends/cache metadata columns, metadata column grouping with collapsible sections, and cache metadata sheet.

import io
import json
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from utils.logger import logger

# Allowed output directory for file exports
_ALLOWED_OUTPUT_DIR = Path(__file__).parent.parent / "outputs"

# Metadata column groups for collapsible sections (Phase 10 Task 12)
_CACHE_METADATA_COLUMNS = [
    "Cache Hit",
    "Cache Key Prefix",
    "Fetched At",
    "Expires At",
    "Cache Provider",
    "Cache TTL Hours",
]

_BM25F_METADATA_COLUMNS = [
    "BM25F Score",
    "BM25F Coverage",
    "Field Contributions",
    "Matched Terms",
]

_SIGNAL_METADATA_COLUMNS = [
    "Title Alignment Score",
    "Title H1 Overlap",
    "Title Rewrite Risk",
    "Content Effort Score",
    "Content Effort Level",
    "Topical Overlap",
    "SimHash64",
]

_TRENDS_METADATA_COLUMNS = [
    "Average Interest",
    "Interest Value",
    "Trends Geo",
    "Trends Timeframe",
    "Trends Provider",
    "Trends Data Confidence",
    "Trends Integrity Warnings",
    "Trends Provider Metadata",
]

# Column group definitions for Excel grouping
_COLUMN_GROUPS = {
    "cache_metadata": {
        "columns": _CACHE_METADATA_COLUMNS,
        "label": "Cache Metadata",
        "default_collapsed": True,
    },
    "bm25f_signals": {
        "columns": _BM25F_METADATA_COLUMNS + _SIGNAL_METADATA_COLUMNS,
        "label": "BM25F & Signals",
        "default_collapsed": False,
    },
    "trends_metadata": {
        "columns": _TRENDS_METADATA_COLUMNS,
        "label": "Trends Data",
        "default_collapsed": False,
    },
}

# CLASS_CONTRACT: ExcelExporter
# Purpose: Export keyword analysis DataFrames as formatted Excel workbooks with multi-sheet and metadata support.
# LINKS: requirements.xml#UC-001, requirements.xml#EXPT-10-01
class ExcelExporter:
    _WRAPPED_KEYWORD_HEADERS = {"Keyword", "Ключевые слова", "Ключові слова"}

    # FUNCTION_CONTRACT: _format_workbook
    # Purpose: Implement the format workbook helper for this module.
    # Input: wb (Workbook), style_rules (Optional[Dict])
    # Output: None
    # Side Effects: Formats workbook in place
    # Business Rules: Applies base formatting then optional style_rules (M3 amendment)
    # Failure Modes: Propagates upstream exceptions and existing fallback paths
    # LINKS: requirements.xml#UC-001, PLAN 09-04 Task 7 (M3 amendment)
    @staticmethod
    def _format_workbook(wb: Workbook, style_rules: Optional[Dict] = None) -> None:
        ws = wb.active

        # Стили заголовков
        header_font: Font = Font(bold=True, color="FFFFFF")
        header_fill: PatternFill = PatternFill(
            start_color="4F81BD", end_color="4F81BD", fill_type="solid"
        )

        # Форматирование заголовков
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
        #    cell.alignment = Alignment(horizontal="center")

        # Автоподбор ширины колонок
        keyword_col_letter: Optional[str] = None
        url_col_letter: Optional[str] = None
        seo_text_col_letter: Optional[str] = None
        page_content_col_letter: Optional[str] = None

        # Поиск специальных колонок
        for cell in ws[1]:
            if cell.value in ExcelExporter._WRAPPED_KEYWORD_HEADERS:
                keyword_col_letter = cell.column_letter
            elif cell.value == "URL":
                url_col_letter = cell.column_letter
            elif cell.value == "SEO Text" or cell.value == "SEO текст":
                seo_text_col_letter = cell.column_letter
            elif cell.value == "Page Content":
                page_content_col_letter = cell.column_letter

        for column in ws.columns:
            max_length: int = 0
            col_letter: str = column[0].column_letter

            # Специальная обработка для колонки с ключевыми словами
            if col_letter == keyword_col_letter:
                ws.column_dimensions[col_letter].width = 40
                for cell in column:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                continue

            # Специальная обработка для URL
            if col_letter == url_col_letter:
                ws.column_dimensions[col_letter].width = 30
                for cell in column:
                    cell.alignment = Alignment(
                        wrap_text=True, horizontal="left", vertical="top"
                    )
                continue

            # Специальная обработка для Page Content
            if col_letter == page_content_col_letter:
                ws.column_dimensions[col_letter].width = 120
                for cell in column:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                continue

            # Специальная обработка для SEO текста
            if col_letter == seo_text_col_letter:
                ws.column_dimensions[col_letter].width = 70
                for cell in column:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                continue

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass

            adjusted_width: int = max_length + 2
            if adjusted_width > 50:
                adjusted_width = 50

            ws.column_dimensions[column[0].column_letter].width = adjusted_width

        # Добавление фильтров
        ws.auto_filter.ref = ws.dimensions

        # Apply custom style rules (M3 amendment)
        if style_rules:
            ExcelExporter._apply_style_rules(ws, style_rules)

    # FUNCTION_CONTRACT: _apply_style_rules
    # Purpose: Apply custom style rules to worksheet cells (M3 amendment)
    # Input: ws (Worksheet), style_rules (Dict)
    # Output: None
    # Side Effects: Modifies cell styles in place
    # Business Rules: style_rules maps column_name to condition/font/border/fill
    # Failure Modes: Silently skips invalid rules
    # LINKS: PLAN 09-04 Task 7 (M3 amendment)
    @staticmethod
    # style_rules format:
    # {
    # "column_name": {
    # "condition": lambda row_dict: bool,
    # "font": {"bold": True, "underline": True},
    # }
    # }
    def _apply_style_rules(ws, style_rules: Dict) -> None:
        # Build column index mapping
        header_map = {}
        for cell in ws[1]:
            if cell.value:
                header_map[cell.value] = cell.column - 1  # 0-based index

        for col_name, rule in style_rules.items():
            target_column = (
                rule.get("target_column")
                or rule.get("column")
                or col_name
            )
            if target_column not in header_map and "_" in col_name:
                target_column = col_name.split("_", 1)[0]

            if target_column not in header_map:
                logger.warning(f"Style rule column '{col_name}' not found in Excel headers")
                continue
            col_idx = header_map[target_column]
            condition = rule.get("condition")
            font_config = rule.get("font", {})

            if not condition or not font_config:
                continue

            # Build font from config
            font_kwargs = {}
            if font_config.get("bold"):
                font_kwargs["bold"] = True
            if font_config.get("underline"):
                font_kwargs["underline"] = "single"

            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx + 1)
                # Build row dict for condition evaluation
                row_dict = {}
                for header, idx in header_map.items():
                    row_dict[header] = ws.cell(row=row_idx, column=idx + 1).value

                try:
                    if condition(row_dict):
                        if font_kwargs:
                            cell.font = Font(**font_kwargs)
                except Exception:
                    pass  # Skip cells that fail condition evaluation

    # FUNCTION_CONTRACT: _validate_export_path
    # Purpose: Validate export path is within allowed directory
    # Input: file_path (str)
    # Output: None
    # Side Effects: Raises ValueError if path is invalid
    # Business Rules: Path must be within outputs/ directory
    # Failure Modes: Raises ValueError with descriptive message
    # LINKS: requirements.xml#UC-001
    @staticmethod
    def _validate_export_path(file_path: str) -> None:
        resolved = Path(file_path).resolve()
        allowed = _ALLOWED_OUTPUT_DIR.resolve()
        try:
            resolved.relative_to(allowed)
        except ValueError:
            raise ValueError(
                f"Export path must be within the outputs/ directory. Got: {file_path}"
            )

    # FUNCTION_CONTRACT: export
    # Purpose: Export DataFrame to Excel file
    # Input: data (pd.DataFrame), file_path (str), style_rules (Optional[Dict])
    # Output: bool
    # Side Effects: Creates Excel file at file_path
    # Business Rules: Validates path, writes data, applies formatting
    # Failure Modes: Returns False on error
    # LINKS: requirements.xml#UC-001
    @staticmethod
    def export(data: pd.DataFrame, file_path: str, style_rules: Optional[Dict] = None) -> bool:
        try:
            ExcelExporter._validate_export_path(file_path)
            # 1. Базовая запись
            data.to_excel(file_path, index=False, engine="openpyxl")

            # 2. Расширенное форматирование
            wb: Workbook = load_workbook(file_path)
            ExcelExporter._format_workbook(wb, style_rules)
            wb.save(file_path)

            logger.info(f"Successfully exported data to {file_path}")
            return True

        except Exception as e:
            logger.error(f"Failed to export Excel: {e}")
            return False

    # FUNCTION_CONTRACT: export_to_buffer
    # Purpose: Export DataFrame to in-memory buffer
    # Input: data (pd.DataFrame), buffer (io.BytesIO), style_rules (Optional[Dict])
    # Output: bool
    # Side Effects: Writes to buffer; applies optional styling (M3 amendment)
    # Business Rules: style_rules is optional and backward-compatible
    # Failure Modes: Returns False on error
    # LINKS: requirements.xml#UC-001
    @staticmethod
    def export_to_buffer(data: pd.DataFrame, buffer: io.BytesIO, style_rules: Optional[Dict] = None) -> bool:
        try:
            data.to_excel(buffer, index=False, engine="openpyxl")
            buffer.seek(0)

            wb: Workbook = load_workbook(buffer)
            ExcelExporter._format_workbook(wb, style_rules)

            buffer.seek(0)
            buffer.truncate()
            wb.save(buffer)
            buffer.seek(0)

            return True

        except Exception as e:
            logger.error(f"Failed to export Excel to buffer: {e}")
            return False

    # FUNCTION_CONTRACT: export_csv
    # Purpose: Export DataFrame to CSV file
    # Input: data (pd.DataFrame), file_path (str)
    # Output: bool
    # Side Effects: Creates CSV file at file_path
    # Business Rules: Uses UTF-8-BOM encoding
    # Failure Modes: Returns False on error
    # LINKS: requirements.xml#UC-001
    @staticmethod
    def export_csv(data: pd.DataFrame, file_path: str) -> bool:
        try:
            ExcelExporter._validate_export_path(file_path)
            Path(file_path).write_bytes(ExcelExporter.export_csv_to_bytes(data))
            logger.info(f"Successfully exported CSV to {file_path}")
            return True
        except Exception as e:
            logger.error(f"Failed to export CSV: {e}")
            return False

    # FUNCTION_CONTRACT: export_csv_to_bytes
    # Purpose: Export DataFrame to CSV bytes
    # Input: data (pd.DataFrame)
    # Output: bytes
    # Side Effects: (none)
    # Business Rules: Uses UTF-8-BOM encoding, CRLF line endings
    # Failure Modes: Returns empty bytes on error
    # LINKS: requirements.xml#UC-001
    @staticmethod
    def export_csv_to_bytes(data: pd.DataFrame) -> bytes:
        csv_body = data.to_csv(index=False, lineterminator="\r\n")
        return csv_body.encode("utf-8-sig")

    # Phase 10 Task 12: Multi-sheet export and metadata grouping methods

    # FUNCTION_CONTRACT: export_multi_sheet
    # Purpose: Export multiple DataFrames to separate sheets with metadata grouping (Phase 10 Task 12)
    # Input: sheets (Dict[str, pd.DataFrame]), file_path (str), style_rules (Optional[Dict]), group_metadata (bool)
    # Output: bool
    # Side Effects: Creates multi-sheet Excel file with grouped metadata columns
    # Business Rules: First sheet is main sheet; supports Trends interest/related/region sheets; groups cache/BM25F/signal metadata
    # Failure Modes: Returns False on export error; logs warnings for invalid sheet names
    # LINKS: requirements.xml#EXPT-10-01, PLAN 10-02 Task 12
    @staticmethod
    def export_multi_sheet(
        sheets: Dict[str, pd.DataFrame],
        file_path: str,
        style_rules: Optional[Dict] = None,
        group_metadata: bool = True,
    ) -> bool:
        """Export multiple DataFrames to separate sheets in one Excel file.

        Args:
            sheets: Dict mapping sheet name to DataFrame
            file_path: Output file path (must be within outputs/)
            style_rules: Optional style rules for main sheet
            group_metadata: Whether to group metadata columns with collapsible sections

        Returns:
            True if export succeeded, False otherwise
        """
        try:
            ExcelExporter._validate_export_path(file_path)

            # Create workbook with first sheet
            sheet_names = list(sheets.keys())
            if not sheet_names:
                logger.warning("No sheets to export")
                return False

            first_name = sheet_names[0]
            first_df = sheets[first_name]

            # Write first sheet
            first_df.to_excel(file_path, sheet_name=first_name, index=False, engine="openpyxl")

            # Append additional sheets
            if len(sheet_names) > 1:
                wb = load_workbook(file_path)
                for sheet_name in sheet_names[1:]:
                    df = sheets[sheet_name]
                    # Sanitize sheet name (Excel max 31 chars, no []:*?/\)
                    safe_name = ExcelExporter._sanitize_sheet_name(sheet_name)
                    if safe_name in wb.sheetnames:
                        safe_name = f"{safe_name[:28]}_{len(wb.sheetnames)}"
                    ws = wb.create_sheet(title=safe_name)
                    ExcelExporter._write_dataframe_to_worksheet(ws, df)

                wb.save(file_path)

            # Format all sheets
            wb = load_workbook(file_path)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                ExcelExporter._format_worksheet(ws, group_metadata=group_metadata)

            # Apply style rules only to first (main) sheet
            if style_rules and wb.sheetnames:
                ExcelExporter._apply_style_rules(wb.active, style_rules)

            wb.save(file_path)

            logger.info(f"Successfully exported multi-sheet Excel to {file_path}")
            return True

        except Exception as e:
            logger.error(f"Failed to export multi-sheet Excel: {e}")
            return False

    # FUNCTION_CONTRACT: export_multi_sheet_to_buffer
    # Purpose: Export multiple sheets to buffer (Phase 10 Task 12)
    # Input: sheets (Dict[str, pd.DataFrame]), buffer (io.BytesIO), style_rules (Optional[Dict]), group_metadata (bool)
    # Output: bool
    # Side Effects: Writes multi-sheet Excel to buffer with grouped metadata
    # Business Rules: Same as export_multi_sheet but to in-memory buffer
    # Failure Modes: Returns False on export error
    # LINKS: requirements.xml#EXPT-10-01, PLAN 10-02 Task 12
    @staticmethod
    def export_multi_sheet_to_buffer(
        sheets: Dict[str, pd.DataFrame],
        buffer: io.BytesIO,
        style_rules: Optional[Dict] = None,
        group_metadata: bool = True,
    ) -> bool:
        """Export multiple DataFrames to separate sheets in a buffer."""
        try:
            if not sheets:
                logger.warning("No sheets to export")
                return False

            # Create new workbook
            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet

            # Add sheets
            for sheet_name, df in sheets.items():
                safe_name = ExcelExporter._sanitize_sheet_name(sheet_name)
                if safe_name in wb.sheetnames:
                    safe_name = f"{safe_name[:28]}_{len(wb.sheetnames)}"
                ws = wb.create_sheet(title=safe_name)
                ExcelExporter._write_dataframe_to_worksheet(ws, df)

            # Format all sheets
            for ws in wb.worksheets:
                ExcelExporter._format_worksheet(ws, group_metadata=group_metadata)

            # Apply style rules to first sheet
            if style_rules and wb.worksheets:
                ExcelExporter._apply_style_rules(wb.worksheets[0], style_rules)

            wb.save(buffer)
            buffer.seek(0)

            return True

        except Exception as e:
            logger.error(f"Failed to export multi-sheet Excel to buffer: {e}")
            return False

    # FUNCTION_CONTRACT: _sanitize_sheet_name
    # Purpose: Sanitize sheet name for Excel compatibility
    # Input: name (str)
    # Output: str
    # Side Effects: (none)
    # Business Rules: Removes invalid chars [];:*?/\, limits to 31 chars
    # Failure Modes: Returns "Sheet" if name is invalid or empty
    # LINKS: PLAN 10-02 Task 12
    @staticmethod
    def _sanitize_sheet_name(name: str) -> str:
        if not name:
            return "Sheet"

        # Remove invalid characters
        invalid_chars = "[]:*?/\\"
        for char in invalid_chars:
            name = name.replace(char, "_")

        # Truncate to 31 chars (Excel limit)
        return name[:31]

    # FUNCTION_CONTRACT: _write_dataframe_to_worksheet
    # Purpose: Write DataFrame to worksheet starting at A1
    # Input: ws (Worksheet), df (pd.DataFrame)
    # Output: None
    # Side Effects: Writes data to worksheet in place
    # Business Rules: Writes header row then data rows
    # Failure Modes: Logs errors for malformed data
    # LINKS: PLAN 10-02 Task 12
    @staticmethod
    def _write_dataframe_to_worksheet(ws: Worksheet, df: pd.DataFrame) -> None:
        # Write header
        for col_idx, column in enumerate(df.columns, 1):
            ws.cell(row=1, column=col_idx, value=str(column))

        # Write data
        for row_idx, row_data in enumerate(df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

    # FUNCTION_CONTRACT: _format_worksheet
    # Purpose: Format individual worksheet with headers and metadata grouping
    # Input: ws (Worksheet), group_metadata (bool)
    # Output: None
    # Side Effects: Formats worksheet in place with optional column grouping
    # Business Rules: Applies header formatting, column width, filter, and metadata grouping
    # Failure Modes: Logs warnings for formatting failures
    # LINKS: PLAN 10-02 Task 12
    @staticmethod
    def _format_worksheet(ws: Worksheet, group_metadata: bool = True) -> None:
        if ws.max_row == 0:
            return

        # Header formatting
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        # Build column map
        header_to_col = {}
        for cell in ws[1]:
            if cell.value:
                header_to_col[str(cell.value)] = cell.column

        # Special column handling
        keyword_col = None
        for header in ExcelExporter._WRAPPED_KEYWORD_HEADERS:
            if header in header_to_col:
                keyword_col = header_to_col[header]
                break

        url_col = header_to_col.get("URL")
        seo_text_col = header_to_col.get("SEO Text") or header_to_col.get("SEO текст")
        page_content_col = header_to_col.get("Page Content")

        # Set column widths and alignment
        for column in ws.columns:
            col_letter = column[0].column_letter
            col_idx = column[0].column

            # Keyword column
            if col_idx == keyword_col:
                ws.column_dimensions[col_letter].width = 40
                for cell in column:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                continue

            # URL column
            if col_idx == url_col:
                ws.column_dimensions[col_letter].width = 30
                for cell in column:
                    cell.alignment = Alignment(wrap_text=True, horizontal="left", vertical="top")
                continue

            # Page Content column
            if col_idx == page_content_col:
                ws.column_dimensions[col_letter].width = 120
                for cell in column:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                continue

            # SEO Text column
            if col_idx == seo_text_col:
                ws.column_dimensions[col_letter].width = 70
                for cell in column:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                continue

            # Auto-fit other columns
            max_len = 0
            for cell in column:
                try:
                    val_len = len(str(cell.value)) if cell.value else 0
                    if val_len > max_len:
                        max_len = val_len
                except Exception:
                    pass

            adjusted_width = min(max_len + 2, 50)
            ws.column_dimensions[col_letter].width = adjusted_width

        # Add filter
        if ws.max_row > 1:
            ws.auto_filter.ref = ws.dimensions

        # Group metadata columns (Phase 10 Task 12)
        if group_metadata:
            ExcelExporter._group_metadata_columns(ws, header_to_col)

    # FUNCTION_CONTRACT: _group_metadata_columns
    # Purpose: Group metadata columns into collapsible sections
    # Input: ws (Worksheet), header_to_col (Dict[str, int])
    # Output: None
    # Side Effects: Creates Excel column groups for metadata sections
    # Business Rules: Groups cache/BM25F/signal/Trends metadata columns; collapses cache metadata by default
    # Failure Modes: Logs warnings for grouping failures
    # LINKS: PLAN 10-02 Task 12
    @staticmethod
    # Creates column outline groups for metadata to control proliferation.
    def _group_metadata_columns(ws: Worksheet, header_to_col: Dict[str, int]) -> None:
        for group_name, group_config in _COLUMN_GROUPS.items():
            columns = group_config["columns"]
            found_cols = []

            for header in columns:
                if header in header_to_col:
                    found_cols.append(header_to_col[header])

            if not found_cols or len(found_cols) < 2:
                continue  # Need at least 2 columns to group

            # Sort column indices
            found_cols.sort()

            # Group from first to last column
            start_col = found_cols[0]
            end_col = found_cols[-1]

            try:
                # Create column group — openpyxl requires column letters, not ints
                start_letter = get_column_letter(start_col)
                end_letter = get_column_letter(end_col)
                ws.column_dimensions.group(start_letter, end_letter, outline_level=1)

                # Collapse if configured
                if group_config.get("default_collapsed", False):
                    ws.column_dimensions.group(start_letter, end_letter, hidden=True)
            except Exception as e:
                logger.warning(f"Failed to group columns {start_col}-{end_col}: {e}")

    # FUNCTION_CONTRACT: add_cache_metadata_columns
    # Purpose: Add cache metadata columns to DataFrame (Phase 10 Task 12)
    # Input: df (pd.DataFrame), cache_metadata (Dict[str, Any])
    # Output: pd.DataFrame
    # Side Effects: Returns new DataFrame with cache columns appended
    # Business Rules: Adds Cache Hit, Cache Key Prefix, Fetched At, Expires At columns
    # Failure Modes: Returns original DataFrame on error
    # LINKS: requirements.xml#EXPT-10-01, PLAN 10-02 Task 12
    @staticmethod
    def add_cache_metadata_columns(
        df: pd.DataFrame,
        cache_metadata: Dict[str, Any],
    ) -> pd.DataFrame:
        """Add cache metadata columns to DataFrame.

        Args:
            df: Source DataFrame
            cache_metadata: Dict with cache_hit (bool), cache_key (str),
                fetched_at (str/datetime), expires_at (str/datetime),
                provider (str), ttl_hours (int)

        Returns:
            DataFrame with cache metadata columns appended
        """
        if df.empty or not cache_metadata:
            return df

        result = df.copy()

        # Cache Hit
        result["Cache Hit"] = "Yes" if cache_metadata.get("cache_hit", False) else "No"

        # Cache Key Prefix (first 16 chars)
        cache_key = cache_metadata.get("cache_key", "")
        result["Cache Key Prefix"] = cache_key[:16] if cache_key else ""

        # Fetched At
        fetched_at = cache_metadata.get("fetched_at") or cache_metadata.get("created_at")
        if fetched_at:
            if isinstance(fetched_at, datetime):
                result["Fetched At"] = fetched_at.strftime("%Y-%m-%d %H:%M")
            else:
                result["Fetched At"] = str(fetched_at)[:19]
        else:
            result["Fetched At"] = ""

        # Expires At
        expires_at = cache_metadata.get("expires_at")
        if expires_at:
            if isinstance(expires_at, datetime):
                result["Expires At"] = expires_at.strftime("%Y-%m-%d %H:%M")
            else:
                result["Expires At"] = str(expires_at)[:19]
        else:
            result["Expires At"] = ""

        # Cache Provider
        result["Cache Provider"] = cache_metadata.get("provider", "")[:50]

        # Cache TTL Hours
        result["Cache TTL Hours"] = cache_metadata.get("ttl_hours", 0)

        return result

    # FUNCTION_CONTRACT: add_bm25f_signal_columns
    # Purpose: Add BM25F and signal analysis columns (Phase 10 Task 12)
    # Input: df (pd.DataFrame), analysis_data (Dict[str, Any])
    # Output: pd.DataFrame
    # Side Effects: Returns new DataFrame with BM25F/signal columns
    # Business Rules: Adds BM25F Score/Coverage, signal scores; maps by keyword if provided
    # Failure Modes: Returns original DataFrame on error
    # LINKS: requirements.xml#EXPT-10-01, PLAN 10-02 Task 12
    @staticmethod
    def add_bm25f_signal_columns(
        df: pd.DataFrame,
        analysis_data: Dict[str, Any],
        keyword_col: str = "Keyword",
    ) -> pd.DataFrame:
        """Add BM25F and signal analysis columns to DataFrame.

        Args:
            df: Source DataFrame
            analysis_data: Dict with bm25f_results (List[BM25FScore]),
                signal_results (Dict with title_alignment, content_effort, etc.)
            keyword_col: Column name to match keywords

        Returns:
            DataFrame with BM25F and signal columns
        """
        if df.empty or not analysis_data:
            return df

        result = df.copy()

        # Initialize new columns
        result["BM25F Score"] = 0.0
        result["BM25F Coverage"] = 0.0
        result["Field Contributions"] = ""
        result["Matched Terms"] = ""
        result["Title Alignment Score"] = 0.0
        result["Title H1 Overlap"] = 0.0
        result["Title Rewrite Risk"] = 0.0
        result["Content Effort Score"] = 0.0
        result["Content Effort Level"] = ""
        result["Topical Overlap"] = 0.0
        result["SimHash64"] = ""

        # BM25F results
        bm25f_results = analysis_data.get("bm25f_results", [])
        if bm25f_results:
            keyword_to_bm25f = {}
            for item in bm25f_results:
                if hasattr(item, "doc_text"):
                    keyword_to_bm25f[item.doc_text] = item
                elif isinstance(item, dict):
                    keyword_to_bm25f.get(item.get("doc_text", ""), item)

            for idx, row in result.iterrows():
                keyword = str(row.get(keyword_col, ""))
                if keyword in keyword_to_bm25f:
                    bm25f = keyword_to_bm25f[keyword]
                    if hasattr(bm25f, "score"):
                        result.at[idx, "BM25F Score"] = round(bm25f.score, 3)
                        result.at[idx, "BM25F Coverage"] = round(bm25f.query_coverage, 3)
                        result.at[idx, "Field Contributions"] = str(bm25f.field_contributions)[:100]
                        result.at[idx, "Matched Terms"] = ", ".join(bm25f.matched_terms)[:50]

        # Signal results
        signals = analysis_data.get("signal_results", {})
        if signals:
            # Title alignment
            title_align = signals.get("title_alignment")
            if title_align and hasattr(title_align, "title_alignment_score"):
                result["Title Alignment Score"] = title_align.title_alignment_score
                result["Title H1 Overlap"] = title_align.title_h1_overlap
                result["Title Rewrite Risk"] = title_align.title_rewrite_risk

            # Content effort
            effort = signals.get("content_effort")
            if effort and hasattr(effort, "effort_score"):
                result["Content Effort Score"] = effort.effort_score
                result["Content Effort Level"] = effort.effort_level

            # Topical overlap
            topical = signals.get("topical_overlap")
            if topical and hasattr(topical, "mean_topical_overlap"):
                result["Topical Overlap"] = topical.mean_topical_overlap

            # SimHash
            simhash = signals.get("simhash")
            if simhash and hasattr(simhash, "simhash64_hex"):
                result["SimHash64"] = simhash.simhash64_hex

        return result

# FUNCTION_CONTRACT: add_trends_columns
# Purpose: Add Google Trends data columns (Phase 10 Task 12)
# Input: df (pd.DataFrame), trends_data (Dict[str, Any])
# Output: pd.DataFrame
# Side Effects: Returns new DataFrame with Trends columns
# Business Rules: Adds Average Interest when new averages are supplied, preserves existing Average Interest otherwise, and adds geo/timeframe/provider columns; maps by keyword
# Failure Modes: Returns original DataFrame on error
# LINKS: requirements.xml#EXPT-10-01, PLAN 10-02 Task 12
    @staticmethod
    def add_trends_columns(
        df: pd.DataFrame,
        trends_data: Dict[str, Any],
        keyword_col: str = "Keyword",
    ) -> pd.DataFrame:
        """Add Google Trends data columns to DataFrame.

        Args:
            df: Source DataFrame
            trends_data: Dict with averages (Dict[str, float]), geo, timeframe, provider
            keyword_col: Column name to match keywords

        Returns:
            DataFrame with Trends columns
        """
        if df.empty or not trends_data:
            return df

        result = df.copy()

        averages = trends_data.get("averages", {}) or {}

        # Initialize columns
        if "Average Interest" not in result.columns:
            result["Average Interest"] = 0.0 if averages else ""
        result["Trends Geo"] = ""
        result["Trends Timeframe"] = ""
        result["Trends Provider"] = ""
        result["Trends Data Confidence"] = ""
        result["Trends Integrity Warnings"] = ""
        result["Trends Provider Metadata"] = ""

        # Map averages by keyword
        if averages:
            for idx, row in result.iterrows():
                keyword = str(row.get(keyword_col, "")).lower()
                # Try exact match first
                if keyword in averages:
                    result.at[idx, "Average Interest"] = round(averages[keyword], 2)
                else:
                    # Try case-insensitive match
                    for trend_kw, value in averages.items():
                        if trend_kw.lower() == keyword:
                            result.at[idx, "Average Interest"] = round(value, 2)
                            break

        # Add metadata columns (same for all rows)
        geo = trends_data.get("geo", "")
        timeframe = trends_data.get("timeframe", "")
        provider = trends_data.get("provider", "")
        data_confidence = trends_data.get("data_confidence", "")
        integrity_warnings = trends_data.get("integrity_warnings", []) or []
        provider_metadata = trends_data.get("provider_metadata", {}) or {}

        result["Trends Geo"] = geo
        result["Trends Timeframe"] = timeframe
        result["Trends Provider"] = provider[:50]
        result["Trends Data Confidence"] = str(data_confidence)[:20]
        result["Trends Integrity Warnings"] = " | ".join(
            str(item) for item in integrity_warnings if str(item).strip()
        )[:200]
        if provider_metadata:
            result["Trends Provider Metadata"] = json.dumps(
                provider_metadata,
                ensure_ascii=False,
                sort_keys=True,
            )[:500]

        return result

    # FUNCTION_CONTRACT: build_trends_provider_metadata_df
    # Purpose: Build a compact metadata DataFrame for Trends exports.
    # Input: trends_data (Dict[str, Any])
    # Output: pd.DataFrame
    # Side Effects: None
    # Business Rules: Preserves provider metadata, confidence, cache metadata, warnings, and caveats for workbook export
    # Failure Modes: Returns an empty frame when no metadata exists
    # LINKS: requirements.xml#EXPT-10-01
    @staticmethod
    def build_trends_provider_metadata_df(trends_data: Dict[str, Any]) -> pd.DataFrame:
        rows: list[dict[str, Any]] = []
        if not trends_data:
            return pd.DataFrame(columns=["Field", "Value"])

        provider = str(trends_data.get("provider", "")).strip()
        data_confidence = str(trends_data.get("data_confidence", "")).strip()
        integrity_warnings = trends_data.get("integrity_warnings", []) or []
        warnings = trends_data.get("warnings", []) or []
        caveats = trends_data.get("caveats", []) or []
        provider_metadata = trends_data.get("provider_metadata", {}) or {}
        cache_metadata = trends_data.get("cache_metadata", {}) or {}

        if provider:
            rows.append({"Field": "Provider", "Value": provider})
        if data_confidence:
            rows.append({"Field": "Data Confidence", "Value": data_confidence})
        if warnings:
            rows.append(
                {
                    "Field": "Warnings",
                    "Value": " | ".join(str(item) for item in warnings if str(item).strip()),
                }
            )
        if integrity_warnings:
            rows.append(
                {
                    "Field": "Integrity Warnings",
                    "Value": " | ".join(
                        str(item) for item in integrity_warnings if str(item).strip()
                    ),
                }
            )
        if provider_metadata:
            rows.append(
                {
                    "Field": "Provider Metadata",
                    "Value": json.dumps(provider_metadata, ensure_ascii=False, sort_keys=True),
                }
            )
        if cache_metadata:
            rows.append(
                {
                    "Field": "Cache Metadata",
                    "Value": json.dumps(cache_metadata, ensure_ascii=False, sort_keys=True),
                }
            )
        for caveat in caveats:
            caveat_text = str(caveat).strip()
            if caveat_text:
                rows.append({"Field": "Caveat", "Value": caveat_text})

        return pd.DataFrame(rows, columns=["Field", "Value"])

    # FUNCTION_CONTRACT: export_merged_report
    # Purpose: Export a merged report combining SERP, Ads, and math analysis data into a single Excel file.
    # Input: output_path (Path), serp_data (pd.DataFrame), ads_data (pd.DataFrame), math_data (pd.DataFrame), report_title (Optional[str])
    # Output: bool
    # Side Effects: Writes Excel file with multiple sheets
    # Business Rules: Creates separate sheets for SERP, Ads, and Math Analysis; includes a Summary sheet
    # Failure Modes: Returns False on error, logs warning
    # LINKS: PLAN 15-03, EXPORT-15-03
    def export_merged_report(
        self,
        output_path: Path,
        serp_data: pd.DataFrame,
        ads_data: pd.DataFrame,
        math_data: pd.DataFrame,
        report_title: Optional[str] = None,
    ) -> bool:
        """Export a merged report combining SERP, Ads, and math analysis data."""
        try:
            self._validate_export_path(output_path)

            wb = Workbook()

            # Remove default sheet
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]

            # Summary sheet
            ws_summary = wb.create_sheet("Summary", 0)
            ws_summary["A1"] = report_title or "Merged SEO Analysis Report"
            ws_summary["A1"].font = Font(bold=True, size=14)
            ws_summary["A3"] = "Generated:"
            ws_summary["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ws_summary["A5"] = "Sections:"
            ws_summary["B5"] = "SERP Analysis, Ads Data, Math Analysis"

            row = 7
            if not serp_data.empty:
                ws_summary[f"A{row}"] = "SERP Results:"
                ws_summary[f"B{row}"] = f"{len(serp_data)} rows"
                row += 1
            if not ads_data.empty:
                ws_summary[f"A{row}"] = "Ads Keywords:"
                ws_summary[f"B{row}"] = f"{len(ads_data)} rows"
                row += 1
            if not math_data.empty:
                ws_summary[f"A{row}"] = "Math Analysis:"
                ws_summary[f"B{row}"] = f"{len(math_data)} rows"

            # SERP Analysis sheet
            if not serp_data.empty:
                ws_serp = wb.create_sheet("SERP Analysis")
                ExcelExporter._write_dataframe_to_worksheet(ws_serp, serp_data)
                ExcelExporter._format_worksheet(ws_serp)

            # Ads Data sheet
            if not ads_data.empty:
                ws_ads = wb.create_sheet("Ads Data")
                ExcelExporter._write_dataframe_to_worksheet(ws_ads, ads_data)
                ExcelExporter._format_worksheet(ws_ads)

            # Math Analysis sheet
            if not math_data.empty:
                ws_math = wb.create_sheet("Math Analysis")
                ExcelExporter._write_dataframe_to_worksheet(ws_math, math_data)
                ExcelExporter._format_worksheet(ws_math)

            # Apply base formatting
            ExcelExporter._format_workbook(wb)

            # Save
            output_path.parent.mkdir(parents=True, exist_ok=True)
            wb.save(str(output_path))
            logger.info(f"Merged report exported to {output_path}")
            return True

        except Exception as e:
            logger.warning(f"Failed to export merged report: {e}")
            return False
